from gtts import gTTS
from base64 import b64encode
import io
import openpyxl


def get_voice(chinese: str):
    tts = gTTS(text=chinese, lang='zh-CN')
    fp = io.BytesIO()
    tts.write_to_fp(fp)
    voice = b64encode(fp.getvalue()).decode('utf-8')
    return voice


def words_from_excel(file):
    validate_header = 'chinese,pinyin,meaning'
    if file:
        workbook = openpyxl.load_workbook(file, data_only=True)
        worksheet = workbook.active
        
        header_row = next(worksheet.iter_rows(min_row=1, values_only=True))
        
        if ','.join(header_row).lower() == validate_header:
            return worksheet.iter_rows(min_row=2, values_only=True)
